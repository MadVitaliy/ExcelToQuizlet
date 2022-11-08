from openpyxl import load_workbook
from openpyxl.utils import *

def FirstNotEmptyRow(sheet):
  for i, row in enumerate(sheet):
    if not all(cell.value is None for cell in row):
      return i + 1
  return -1
  
  
def LastNotEmptyRow(sheet):
  for i, row in reversed(list(enumerate(sheet))):
    if not all(cell.value is None for cell in row):
      return i + 1
  return -1
  
def NumberOfNotEmptyRow(sheet):
  counter = int(0);
  for i, row in enumerate(sheet):
    if not all(cell.value is None for cell in row):
      counter += 1
  return counter

def ReadColumns(sheet, colum_inds):
  exel_column_inds = [get_column_letter(ind) for ind in colum_inds]
  
  print(exel_column_inds)


def ReadExcel(path):
  wb = load_workbook(path)
  ws = wb.active
  
  first_row_ind = FirstNotEmptyRow(ws)
  last_row_ind = LastNotEmptyRow(ws)
  number_of_not_empt = NumberOfNotEmptyRow(ws)

  #reading table header
  header_row_ind = first_row_ind
  titles_inds = [i+1 for i, cell in enumerate(ws[header_row_ind]) if cell.value is not None]
  titles = {}
  for i in titles_inds:
    cell = ws.cell(row=header_row_ind, column=i)
    value = cell.value.strip().lower()
    titles[i] = value 

  #reading all content
  data =[]
  for row_ind in range(header_row_ind+1, last_row_ind+1):
    row_data = {}
    for column_ind in titles_inds:
      cell = ws.cell(row=row_ind, column=column_ind)
      #print(cell)
      value = cell.value
      row_data[titles[column_ind]] = value
    data.append(row_data)
  
  return [val for key,val in titles.items()], data
